#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Example client for VHELIBS Web's REST API.

What this script does:

  1. Starts `run.py` (the Flask server) as a separate process.
  2. Waits until the server actually responds before sending anything.
  3. Submits the given PDB IDs in SMALL BATCHES, one after another, instead
     of putting hundreds of IDs into a single /api/analyse request.
  4. For each batch, polls GET /api/status/<job_id> until it's done, turns
     each structure's result into spreadsheet rows, and discards the raw
     analysis JSON (atom coordinates, density boxes/atoms, etc.) as soon as
     those rows have been extracted.
  5. Once every batch has been processed, writes a two-sheet .xlsx
     ("Ligands" + "Parameters") in the exact same layout produced by the
     Results tab's own "Export (.xlsx)" button.

This uses less memory than the GUI export because each batch is fetched,
immediately flattened into a handful of short text values per ligand, and
thrown away. What's kept in memory for the whole run is just that flat list
of spreadsheet rows — typically a few hundred bytes per ligand — not the
full analysis payload.

Requirements: `pip install requests openpyxl`

Usage:
    python vhelibs_batch_client.py \\
        --repo-root /path/to/vhelibs-web \\
        --pdbids 1cbs 3dzu 4hhb 1a28 P00734 \\
        --batch-size 3 \\
        --output results.xlsx
"""
import argparse
import itertools
import subprocess
import sys
import time
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Spreadsheet layout 
# ---------------------------------------------------------------------------

XLSX_NOT_AVAILABLE = "Not Available"
XLSX_MAX_COL_WIDTH = 30
XLSX_MIN_COL_WIDTH = 8

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
NOWRAP_ALIGNMENT = Alignment(wrap_text=False, vertical="top")

LIGANDS_HEADER = [
    "UniProt ID", "PDB ID", "Ligand ID", "Ligand Name",
    "Ligand Class.", "BS Class.", "BS Residues",
    "R-free", "R-work", "Rejected Molecules",
    "EDM Available",
]

LIGANDS_NOWRAP_COLUMNS = {"BS Residues"}

PARAM_LABELS = [
    ("pdbids",           "PDB IDs / UniProt IDs (as entered)"),
    ("rsr_upper",        "RSR upper threshold (Bad above this)"),
    ("rsr_lower",        "RSR lower threshold (Good below this)"),
    ("rscc_min",         "RSCC minimum"),
    ("rfree_max",        "R-free maximum"),
    ("occupancy_min",    "Occupancy minimum"),
    ("tolerance",        "Tolerance"),
    ("distance",         "Binding site distance (A)"),
    ("use_pdb_redo",     "Use PDB-REDO structures"),
    ("check_owab",       "Check OWAB"),
    ("owab_max",         "OWAB maximum"),
    ("check_resolution", "Check resolution"),
    ("resolution_max",   "Resolution maximum (A)"),
    ("use_rdiff",        "Use R-diff filter"),
    ("rdiff_max",        "R-diff maximum"),
    ("use_dpi",          "Use DPI filter"),
    ("dpi_max",          "DPI maximum"),
    ("use_cache",        "Use cached downloads"),
]

QUALITY_LETTER = {"Bad": "B", "Dubious": "D", "Good": "G"}
QUALITY_SEVERITY = {"Bad": 0, "Dubious": 1, "Good": 2}


def or_na(value):
    if value is None or value == "":
        return XLSX_NOT_AVAILABLE
    return value


def fmt_num(value):
    return value if isinstance(value, (int, float)) else None


def rejected_residues_text(rejected):
    keys = list((rejected or {}).keys())
    return "; ".join(keys)


def format_ligand_names(ligand):
    names = ligand.get("ligand_names") or []
    if not names:
        return ""
    return "; ".join(n or XLSX_NOT_AVAILABLE for n in names)


def format_binding_site_residues(ligand):
    """Binding-site residues annotated with their quality letter, worst first
    (a residue absent from residue_qualities was never flagged, so it's Good
    by construction)"""
    residues = ligand.get("binding_site_residues") or []
    if not residues:
        return ""
    qualities = ligand.get("residue_qualities") or {}
    annotated = [(res, qualities.get(res, "Good")) for res in residues]
    annotated.sort(key=lambda pair: QUALITY_SEVERITY[pair[1]])
    return "; ".join(f"{res} ({QUALITY_LETTER[quality]})" for res, quality in annotated)


def format_param_value(value):
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    if value is None:
        return ""
    return value


# ---------------------------------------------------------------------------
# 1) Starting and stopping the server
# ---------------------------------------------------------------------------

def start_server(repo_root, host, port, cache_dir=None):
    """Launch `run.py` as a separate process and return the Popen handle."""
    cmd = [sys.executable, "run.py", "--host", host, "--port", str(port), "--no-browser"]
    if cache_dir:
        cmd += ["--cache-dir", cache_dir]
    print(f"[server] starting: {' '.join(cmd)} (cwd={repo_root})")
    return subprocess.Popen(cmd, cwd=repo_root)


def stop_server(proc):
    if proc.poll() is not None:
        return
    print("[server] stopping...")
    proc.terminate()
    try:
        proc.wait(timeout=10)
    except subprocess.TimeoutExpired:
        proc.kill()
        proc.wait()


def wait_for_server(base_url, timeout=60.0, interval=0.5):
    """Poll `base_url` until the server responds, or until `timeout` seconds pass."""
    print(f"[server] waiting for {base_url} to become available...")
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            requests.get(base_url, timeout=2)
            print("[server] ready.")
            return True
        except requests.exceptions.ConnectionError:
            time.sleep(interval)
    return False


# ---------------------------------------------------------------------------
# 2) Submitting analyses in small batches
# ---------------------------------------------------------------------------

def chunked(seq, size):
    """Split `seq` into chunks of at most `size` elements."""
    it = iter(seq)
    while True:
        chunk = list(itertools.islice(it, size))
        if not chunk:
            return
        yield chunk


def submit_batch(base_url, pdbids_batch, **analysis_params):
    """POST /api/analyse for one small batch of PDB IDs. Returns (job_id, total)."""
    payload = {"pdbids": " ".join(pdbids_batch)}
    payload.update(analysis_params)
    resp = requests.post(f"{base_url}/api/analyse", json=payload, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    return data["job_id"], data.get("total")


def poll_job(base_url, job_id, poll_interval=1.5, timeout=1800):
    """GET /api/status/<job_id> repeatedly until status == 'done'."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        resp = requests.get(f"{base_url}/api/status/{job_id}", timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if data["status"] == "done":
            return data
        print(f"    ... progress {data.get('progress')}/{data.get('total')}")
        time.sleep(poll_interval)
    raise TimeoutError(f"Job {job_id} did not finish within {timeout}s")


def resolve_edm_availability(base_url, struct, use_pdb_redo):
    """Whether the electron-density MAP file is available for one structure,
    without downloading it.

    Returns True / False / None (None = could not be determined).
    """
    if struct.get("error"):
        return None
    if use_pdb_redo:
        # PDB-REDO's map-maker computes the map on request rather than
        # serving a static file; a successful PDB-REDO analysis already
        # implies the map is obtainable, so no extra request is needed.
        return True
    pdbid = (struct.get("pdbid") or "").lower()
    try:
        resp = requests.get(f"{base_url}/api/edm-exists/{pdbid}", timeout=15)
        resp.raise_for_status()
        return bool(resp.json().get("exists"))
    except requests.RequestException:
        return None


# ---------------------------------------------------------------------------
# 3) Flattening one /api/status result into Ligands-sheet rows
# ---------------------------------------------------------------------------

def rows_for_structure(struct, edm_available):
    """One or more flat rows for one entry of `results`."""
    complex_id = (struct.get("pdbid") or "").upper()
    uniprot = or_na(struct.get("uniprot"))

    if struct.get("error"):
        # The backend never got past the initial fetch/parse for this
        # entry, so nothing beyond the UniProt tag and complex ID exists.
        yield [uniprot, complex_id] + [XLSX_NOT_AVAILABLE] * 9
        return

    struc_dict = struct.get("struc_dict") or {}
    r_free = or_na(fmt_num(struc_dict.get("rFree")))
    r_work = or_na(fmt_num(struc_dict.get("rWork")))
    rejected_text = rejected_residues_text(struct.get("rejected"))
    map_avail = (
        "Yes" if edm_available is True else
        "No" if edm_available is False else
        XLSX_NOT_AVAILABLE
    )

    ligands = struct.get("ligands") or []
    if not ligands:
        # Analysed successfully but nothing qualified as a ligand (e.g.
        # everything present was blacklisted).
        yield [uniprot, complex_id] + [XLSX_NOT_AVAILABLE] * 5 + [r_free, r_work, rejected_text, map_avail]
        return

    for ligand in ligands:
        yield [
            uniprot,
            complex_id,
            or_na("; ".join(ligand.get("ligand_residues") or [])),
            or_na(format_ligand_names(ligand)),
            or_na(ligand.get("ligand_quality")),
            or_na(ligand.get("binding_site_quality")),
            or_na(format_binding_site_residues(ligand)),
            r_free,
            r_work,
            rejected_text,
            map_avail,
        ]


# ---------------------------------------------------------------------------
# 4) Writing the two-sheet .xlsx (same layout as the GUI's Export button)
# ---------------------------------------------------------------------------

def compute_column_widths(rows):
    """Longest cell value per column, capped at XLSX_MAX_COL_WIDTH."""
    col_count = max((len(row) for row in rows), default=0)
    widths = [XLSX_MIN_COL_WIDTH] * col_count
    for row in rows:
        for i, cell in enumerate(row):
            length = len(str(cell)) if cell is not None else 0
            widths[i] = min(XLSX_MAX_COL_WIDTH, max(widths[i], length))
    return widths


def add_wrapped_sheet(workbook, name, rows, nowrap_columns=()):
    """Add `rows` (list-of-lists, first row = header) as a new worksheet,
    with column widths capped at XLSX_MAX_COL_WIDTH and "Wrap Text"
    formatting throughout.
    """
    ws = workbook.create_sheet(name)

    for i, width in enumerate(compute_column_widths(rows)):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    header = rows[0] if rows else []
    nowrap_indices = {header.index(col) for col in nowrap_columns if col in header}

    for row in rows:
        cells = []
        for i, value in enumerate(row):
            cell = WriteOnlyCell(ws, value=value)
            cell.alignment = NOWRAP_ALIGNMENT if i in nowrap_indices else WRAP_ALIGNMENT
            cells.append(cell)
        ws.append(cells)

    return ws


def build_parameters_rows(cfg):
    rows = [["Parameter", "Value"]]
    for key, label in PARAM_LABELS:
        rows.append([label, format_param_value(cfg.get(key))])

    blacklist = cfg.get("blacklist") or {}
    disabled = blacklist.get("disabled") or []
    custom = blacklist.get("custom") or []
    rows.append(["Blacklist: disabled default entries", ", ".join(disabled)])
    rows.append(["Blacklist: custom entries added",
                 ", ".join(f"{e['code']} ({e['name']})" for e in custom)])
    rows.append(["Blacklist: default list replaced by uploaded file",
                 "Yes" if blacklist.get("replace") else "No"])
    export_date = datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")
    rows.append(["Export date", export_date])
    return rows


# ---------------------------------------------------------------------------
# 5) Orchestration
# ---------------------------------------------------------------------------

def run(args):
    base_url = f"http://{args.host}:{args.port}"
    proc = start_server(args.repo_root, args.host, args.port, args.cache_dir)
    try:
        if not wait_for_server(base_url, timeout=args.startup_timeout):
            raise RuntimeError("Server did not respond in time; check --repo-root")

        analysis_params = dict(
            rsr_upper=args.rsr_upper, rsr_lower=args.rsr_lower,
            rscc_min=args.rscc_min, rfree_max=args.rfree_max,
            occupancy_min=args.occupancy_min, tolerance=args.tolerance,
            distance=args.distance, use_pdb_redo=args.use_pdb_redo,
            check_owab=args.check_owab, owab_max=args.owab_max,
            check_resolution=args.check_resolution, resolution_max=args.resolution_max,
            use_rdiff=args.use_rdiff, rdiff_max=args.rdiff_max,
            use_dpi=args.use_dpi, dpi_max=args.dpi_max,
            use_cache=not args.no_cache,
        )

        # Only the flattened rows are kept for the whole run.
        ligand_rows = []

        total_batches = -(-len(args.pdbids) // args.batch_size)  # ceil
        for i, batch in enumerate(chunked(args.pdbids, args.batch_size), start=1):
            print(f"[batch {i}/{total_batches}] submitting {batch}")
            job_id, total = submit_batch(base_url, batch, **analysis_params)
            print(f"    job_id={job_id} total={total}")

            result = poll_job(base_url, job_id, poll_interval=args.poll_interval)

            for struct in result["results"]:
                edm_available = resolve_edm_availability(base_url, struct, args.use_pdb_redo)
                ligand_rows.extend(rows_for_structure(struct, edm_available))

            for warning in result.get("warnings") or []:
                print(f"    [warning] {warning}")

        cfg_for_sheet = dict(analysis_params)
        cfg_for_sheet["pdbids"] = " ".join(args.pdbids)
        cfg_for_sheet["blacklist"] = {"disabled": [], "custom": [], "replace": None}

        wb = Workbook(write_only=True)
        # Ligands sheet first, Parameters second.
        add_wrapped_sheet(wb, "Ligands", [LIGANDS_HEADER] + ligand_rows, LIGANDS_NOWRAP_COLUMNS)
        add_wrapped_sheet(wb, "Parameters", build_parameters_rows(cfg_for_sheet))
        wb.save(args.output)
        print(f"[done] results saved to {args.output}")
    finally:
        stop_server(proc)


def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--repo-root", required=True, help="Directory containing run.py")
    p.add_argument("--host", default="127.0.0.1")
    p.add_argument("--port", type=int, default=8000)
    p.add_argument("--cache-dir", default=None)
    p.add_argument("--startup-timeout", type=float, default=60.0)

    p.add_argument("--pdbids", nargs="+", required=True,
                    help="PDB IDs and/or UniProt IDs, e.g. 1cbs 3dzu 4hhb P00734")
    p.add_argument("--batch-size", type=int, default=5,
                    help="How many IDs to send per /api/analyse request")
    p.add_argument("--poll-interval", type=float, default=1.5,
                    help="Seconds between polls of /api/status/<job_id>")

    p.add_argument("--rsr-upper", type=float, default=0.4, dest="rsr_upper")
    p.add_argument("--rsr-lower", type=float, default=0.24, dest="rsr_lower")
    p.add_argument("--rscc-min", type=float, default=0.9, dest="rscc_min")
    p.add_argument("--rfree-max", type=float, default=1.0, dest="rfree_max")
    p.add_argument("--occupancy-min", type=float, default=1.0, dest="occupancy_min")
    p.add_argument("--tolerance", type=int, default=2)
    p.add_argument("--distance", type=float, default=4.5)
    p.add_argument("--use-pdb-redo", action="store_true", dest="use_pdb_redo")

    p.add_argument("--check-owab", action="store_true", dest="check_owab")
    p.add_argument("--owab-max", type=float, default=50, dest="owab_max")
    p.add_argument("--check-resolution", action="store_true", dest="check_resolution")
    p.add_argument("--resolution-max", type=float, default=3.5, dest="resolution_max")
    p.add_argument("--use-rdiff", action="store_true", dest="use_rdiff")
    p.add_argument("--rdiff-max", type=float, default=0.05, dest="rdiff_max")
    p.add_argument("--use-dpi", action="store_true", dest="use_dpi")
    p.add_argument("--dpi-max", type=float, default=0.42, dest="dpi_max")
    p.add_argument("--no-cache", action="store_true",
                    help="Disable server-side reuse of cached downloads for this run")

    p.add_argument("--output", default="vhelibs_results.xlsx")
    return p.parse_args()


if __name__ == "__main__":
    run(parse_args())
